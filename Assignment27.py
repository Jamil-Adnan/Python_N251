from openpyxl import load_workbook
from openpyxl.styles import Alignment

assignment = load_workbook("Assignment.xlsx")
sheet = assignment.active

for i in range (2, sheet.max_row + 1):
    age = sheet.cell(row = i, column = 2).value
    if age % 2 == 0:
        sheet.cell(row = i, column = 4).value = "Even"
        sheet.cell(row = i, column = 4).alignment = Alignment(horizontal= "center", vertical= "center")
    else:
        sheet.cell(row = i, column = 4).value = "Odd"
        sheet.cell(row = i, column = 4).alignment = Alignment(horizontal= "center", vertical= "center")

# for i in range (2,8):    
#     sheet.cell(row = i, column = 4).value = "Even" if (sheet.cell(row = i, column = 2).value) % 2 == 0 else "Odd" 
    
assignment.save("Assignment.xlsx")
assignment.close()