from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment   


myfile = load_workbook("myfile.xlsx")
sheet = myfile.active

#sheet.append(["Akkas", 36,"M", 5])

# file = open ("Class27.txt", "a")
# for i in sheet.iter_rows(values_only=True):
#     file.write(f"{i}\n")    #file.write(str(i)+"\n")
# # for i in sheet.iter_rows(values_only=True):
# #     print(i)

# myfile.save("myfile.xlsx")
# myfile.close()
# file.close()

file = open ("Class27test.txt", "w+")
for i in sheet.iter_rows(values_only=True):
    file.write(f"{i}\n")    #file.write(str(i)+"\n")

myfile.close()
file.close()